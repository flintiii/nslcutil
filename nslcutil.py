#!/usr/bin/python
"""nclsutil.py

Usage:  nslcutil.py (-f | --2xlsx    ) <ifile> <ofile>
        nslcutil.py (-c | --testcsv  ) <ifile> 
        nslcutil.py (-s | --tespace  ) <ifile> <logfile>    
        nslcutil.py (-t | --testtab  ) <ifile> 
        nslcutil.py (-a | --formh&f  ) <ifile> <ofile> [--tab|--csv]
        nslcutil.py (-m | --transmit ) <ifile>
        nslcutil.py (-h | --details  )
        nslcutil.py --version

Options:
  -h --help     Show this screen.
  --version     Show version.

"""
import re
import os
import csv
import sys
import stat
import subprocess
import signal
import datetime
import io
import string
import logging
# To implement logging see SOURCE:https://docs.python.org/2/howto/logging.html
#W 
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
#
from docopt import docopt
#
from subprocess import call
#
# Boilerplate imports for Python 2 / Python 3 mutual compatiibility
#
# from __future__ import print_function  # Make print a function
#W from six.moves  import input           # Use raw_input when I say input
from os.path    import expanduser      # Cross-platform home directory finder
#
# This slug is based upon the work of Kevin Cole... 
__author__     = "Flint"
__copyright__  = "Copyright 2016, Goddard College (02/01/2016)"
__credits__    = ["Flint"]  # Author and bug reporters
__license__    = "GPL I"
__version__    = "1.00"
__maintainer__ = "Flint"
__email__      = "flint@flint.com"
__status__     = "Production"  # "Prototype", "Development" or "Production"
__appname__    = "NSLC file utililty"
#
# Load Globals
# global alfadefs
#
# Load landers
# Column for Excel Format Index 
length = 0 
start = 1 
stop = 2
reqd = 3
dtype = 4
fieldname = 6
count = 7
spreadsheet = "example.xlsx"
# csvfile = "example.csv"
# csvfile = "bad.example.csv"
comprog = "bin/sf.sh"
# comprog = "bin/hw.sh"
#
from nslcobjects import *

#################### Function Land ####################
#
def poplst():
    """
    Takes ordered directory objects and makes them into a global list of methods
    note the horror of synching this up requires a slug value alfafefs[0]
    """
    global alfadefs  #makes this dictionary global
    alfadefs = {'key':'value'}
    alfadefs[0] = {'key':'value'}
    key=1
    while key <= 109:
        value = dir(D1)[key-1]
        alfadefs[key] = "D1."+value+"()"
        key=key+1
    # test this with:
    # eval(alfadefs[79]) 
    # len(eval(alfadefs[7]).split("," ))
    # print alfadefs[n]+"\t"+eval(alfadefs[n]) 
    # return eval(alfadefs[n]) 
    # return alfadefs[n] 
    # print alfadefs[n][3:-2]+"\t"+eval(alfadefs[n])
    # print (alfadefs[n][3:-2],eval(alfadefs[n])) 
    # return alfadefs[n] 
    # return eval(alfadefs[n])
    #D print " alfadefs is ready to go..."

def lneval(file,symbol):
	"""
	Opens file, skips header.
	returns each line
	"""
	#D file = "example.csv"
	print "Testing " + file + " with " + symbol + " delimiter"
	of = open(file, 'r')
	for line in of:
		if "D1" in line:
			ceval(line,symbol)
	
def ceval(input,symbol):
	"""
	This is a general purpose evaluator for comma seperated value (CSV) files (CSV).  
	How it works is that it knows the following:
	field number (must be same as index number below)
	1 - length
	2 - start
	3 - stop
	4 - required values "R", "O", or "C"
	5 - Type; A, AN, N
	6 - excell column value
	7 - index number (must be same as field number above)
	8 - Fieldname
	It uses this information to check the current element and returns a pass/fail,
	based upon the prepositional logic in "THE HEART OF THE MATTER".
	There is a sample element called oddly enough, sample and it can be adjusted 
	"""
	# this sample contains a - and a . in the city...
	# sample="0,D1,975464835,George,,Washington,,,,F,,804 West 15th Street,,S Signal-Mt.,MS,39440,US,20160626,19670526,20150626,20151123,s,N,,,,,,,,,,,,,,,,,,,,,,,,,,Y,231302,2010,05,104000,W,039000,20130628,N,F,20150626,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,"
	status = " no decision, a bug"
	# here we split on a comma...
	if "comma" in symbol:
		sample = "0,"+input
		mylist = sample.split(',')
		# wish everyone started at one or zero..
	# here we split on a tab...
	if  "tab"  in symbol:
		sample = "0\t"+input
		mylist = sample.split('\t')
		# wish everyone started at one or zero..
	for n in range(1,109):
	# for n in range(1,10):
		element = mylist[n]
		name = alfadefs[n][6:-2]
		criteria = eval(alfadefs[n]).split(',')
		aoran = criteria[4].strip()
		# 
		# THE HEART OF THE MATTER
		if criteria[3] == "O" or criteria[3] == "C" : status = "  o OK"
		elif aoran == "AN" and  element.replace(" ", "").isalnum() and  len(element) <= int(criteria[0]): status = " an OK" 
		elif "SSN" in name and aoran ==  "N" and  element.isdigit() and len(element) == int(criteria[0]): status = "ssn OK" 
		elif "Date" in name and aoran ==  "N" and  element.isdigit() and len(element) == int(criteria[0]): status = "date OK" 
		elif not "SSN" and not "Date" in name and aoran ==  "N" and  element.isdigit() and len(element) <= int(criteria[0]): status = "  n OK" 
		elif "SSN" in name and "NO" in element: status = "ssn OK" 
		elif "Filler" in name: status = "ssn OK" 
		# elif "City" in name and  element.replace(" ", "").isalnum() and  len(element) <= int(criteria[0]): status = " an OK" 
		elif "City" in name and  \
			element.replace(" ", "").replace("-", "").replace(".", "").isalnum() and  \
			len(element) <= int(criteria[0]): status = " an OK" 
		elif "State" in name and  \
			"NO" in mylist[2] or not "US" in mylist[16] : status = " an OK"
		elif "Last" in name and  \
			element.replace(" ", "").replace("-", "").replace(".", "").replace("'", "").isalnum() and  \
			len(element) <= int(criteria[0]): status = " an OK" 
		elif "First" in name and  \
			element.replace(" ", "").replace("-", "").replace(".", "").replace("'", "").isalnum() and  \
			len(element) <= int(criteria[0]): status = " an OK" 
		elif "Street" in name and  len(element) <= int(criteria[0]): status = " an OK" 
		elif aoran ==  "A" and element.isalpha() and len(element) <= int(criteria[0]): status = "  a OK"
		else:status = " Wrong"
		#
		# testers... 
		# print mylist[5]
		# if "SSN" in name: print str(n) + "\t" + str(len(element)) +"\t"+eval(alfadefs	[n]) + "\t" + str(status) + "\t" + aoran + "\t" + alfadefs[n][6:-2]
		# if "Date" in name: print mylist[3] + " " + mylist[5] + "\t" + str(n) + "\t" + mylist[n] + "\t" + str(len(element)) +"\t"+eval(alfadefs[n]) + "\t" + str(status) + "\t" + aoran  + "\t" + name 
		'bin/sf.sh'# if not "OK" in status: print str(n) + "\t" + element + "\t" + str(len(element)) +"\t"+criteria[0] + "\t" + str(status) + "\t" + aoran + "\t" + alfadefs[n][6:-2]
		# if not "OK" in status: print mylist[5] + "\t" + print str(n) + "\t" + str(len(element)) +"\t"+eval(alfadefs[n]) + "\t" + str(status) + "\t" + aoran + "\t" + alfadefs[n][6:-2]
		# if not "OK" in status: print mylist[5] + "\t" + str(n) + "\t" + element + "\t" + str(len(element)) +"\t"+eval(alfadefs[n]) + "\t" + str(status) + "\t" + aoran + "\t" + alfadefs[n][6:-2]
		# if not "OK" in status: print mylist[5] + "\t" + str(n) + "\t" + element + "\t" + element.replace(" ", "").replace("-", "").replace(".", "")
		if not "OK" in status: print mylist[3] + " " + mylist[5] + "\t" + str(n) + "\t" + mylist[n] + "\t" + str(len(element)) +"\t"+eval(alfadefs[n]) + "\t" + str(status) + "\t" + aoran  + "\t" + name 
#
#
def tneval():
	"""
	Just outputs the evaluator.  It prints out what the evaluator uses for criteria
	This is a general purpose evaluator.  How it works is that it knows the following:
	field number (must be same as index number below)
	1 - length
	2 - start
	3 - stop
	#4 - required values "R", "O", or "C"
	5 - Type; A, AN, N
	6 - excell column value
	7 - index number (must be same as field number above)
	8 - Fieldname
	It uses this information to check the current element and returns a pass/fail.
	There is a sample element called oddly enough, sample
	"""
	for n in range(1,110):
		name = alfadefs[n][6:-2]
		# criteria = eval(alfadefs[n]).split(',')
		# print str(criteria) + "\t" + element
		print str(n) +"\t"+eval(alfadefs[n]) + "\t\t" + alfadefs[n][6:-2]
# 
def csv2xlsx(inp,out):
	# Add csv file to an xlsx
	# SOURCE:http://stackoverflow.com/questions/24971556/openpyxl-python-writing-csv-to-excel-gives-number-formatted-as-text
	#
	csvfile = inp
	wb = openpyxl.Workbook()
	ws = wb.active
	with open(csvfile, 'rb') as f:
	    reader = csv.reader(f)
	    for r, row in enumerate(reader, start=1):
		for c, val in enumerate(row, start=1):
		    ws.cell(row=r, column=c).value = val
	wb.save(out)
#
def fhandf(inp,out,symbol):
	print "Welcome to header and footer"
	print inp + "\t" + out  + "\t" + symbol
	ffooter()  		# creates global dictionary of accumulated values 
	with open( out, mode='w') as fout:
		fout.write(fhead(symbol))
		of = open(inp, 'r')
		for line in of:
			if "D1" in line:
				fout.write(line)
		print "There will be a footer at end of " + out +" with " + symbol + " delimiter"
		fout.write(feval(inp,symbol))	# evaluates and writes the footer


def fhead(symbol):
	# symbol = "tab"
	# symbol = "csv"
	today = datetime.datetime.today()
	ddate = today.strftime('%Y%m%d')
	print "Welcome to header - First, a few questions. Just hit enter for default."
	default = "003686"
	scode = str(raw_input("Enter the school code : (default: %s):\n" % default ) or default)
	brcode = "00"
	default = "Spring 2016"
	adyear = str(raw_input("Enter the Academic Term : (default: %s):\n" % default ) or default)
	rptflg = "Y"
	default = str(today.strftime('%Y%m%d'))
	cdate = str(raw_input("Enter certification date yyyymmmdd : (default: %s):\n" % default ) or default)
 	rptlvl = "F"
	os.system('cls' if os.name == 'nt' else 'clear')
	print "This will be in your header, Please check it over."
	print "scode  = " + scode + "\t\t\t",
	print "brcode = " + brcode
	print "adyear = " + adyear + "\t\t",
	print "rptflg = " + rptflg
	print "cdate  = " + str(cdate) + "\t\t",
	print "rptlvl = " + rptlvl
	# A3line = "A3" + d + scode + d + brcode + d + adyear + d + rptflg + d + cdate  + d + rptflg  + d + "Filler is START START START "
	if "tab" in symbol: A3line = "A3" + "\t" + scode + "\t" + \
		brcode + "\t" + adyear + "\t" + rptflg + "\t" + cdate  + "\t" + rptlvl  + "\t" + "Filler is START START START " + "\n"
	if "csv" in symbol: A3line = "A3" + "," + scode + "," + \
		brcode + "," + adyear + "," + rptflg + "," + cdate  + "," + rptlvl  + "," + "Filler is START START START " + "\n"
	# print A3line
	return A3line
#
#
def ffooter():
	"""
	Creates a dictionary of accumulated values.  Usefull? Meh...
	"""
	global facum  #makes this dictionary global
	facum = {'key':'value'}
	key=0
	while key <= 10:
		value = dir(T1)[key-1]
		facum[key] = value
		key=key+1
	#D print " facum is ready to go..

#
def feval(file,symbol):
	"""
	Opens file, skips header.
	returns each line
	"""
	# symbol = "csv"
	# symbol = "tab"
	n = 0
	#D 	file = "example"+"."+symbol
	Fs = Qs = Hs = Ls = Ws = Gs = As = Xs = Ds = Total = 0
	of = open(file, 'r')
	for line in of:
		if "D1" in line:
			# print line
			n += 1
			# ceval(line,symbol)
			#
			# print symbol
			# here we split on a comma...
			if "csv" in symbol:
				mylist = line.split(',')
			# here we split on a tab...
			if "tab" in symbol:
				mylist = line.split('\t')
			#
			# test for proper number of elements
			if len(mylist) != 109: print "Line " + str(n) + " has " +str(len(mylist))+ " elements.  This is Wrong!"
			# print str(len(mylist))
			# if scl 
			# if "F" in mylist[8]: Fs += 1
			if "F" in mylist[8]: Fs += 1
			if "Q" in mylist[8]: Qs += 1
			if "H" in mylist[8]: Hs += 1
			if "L" in mylist[8]: Ls += 1
			if "W" in mylist[8]: Ws += 1
			if "G" in mylist[8]: Gs += 1
			if "A" in mylist[8]: As += 1
			if "X" in mylist[8]: Xs += 1
			if "D" in mylist[8]: Ds += 1
			Total += 1
			# print mylist[8] + " " + str(Fs) + " " + str(Qs) + " " + str(Hs) + " " + str(Ls) + " " + \
			#	str(Ws) + " " + str(Gs) + " " + str(As) + " " + str(Xs) + " " + str(Ds) + " " + str(Total)
		facum[1] = 'Total Number = ' + str(Total)
		facum[2] = 'Number_of_A = ' + str(As)
		facum[3] = 'Number_of_D = ' + str(Ds)
		facum[4] = 'Number_of_F = ' + str(Fs)
		facum[5] = 'Number_of_G = ' + str(Gs)
		facum[6] = 'Number_of_H = ' + str(Hs)
		facum[7] = 'Number_of_L = ' + str(Ls)
		facum[8] = 'Number_of_Q = ' + str(Qs)
		facum[9] = 'Number_of_W = ' + str(Ws)
		facum[10] = 'Number_of_X = ' + str(Xs)	
	#
	# print symbol
	if "csv" in symbol: T1line = "T1" + "," + str(Fs) + "," + str(Qs) + "," + str(Hs) + "," + \
		str(Ls) + "," + str(Ws) + "," + str(Gs) + "," + str(As) + "," + \
		str(Xs) + "," + str(Ds) + "," + str(Total) + "," + "Filler is END END END \n"
	if "tab" in symbol: T1line = "T1" + "\t" + str(Fs) + "\t" + str(Qs) + "\t" + str(Hs) + "\t" + \
		str(Ls) + "\t" + str(Ws) + "\t" + str(Gs) + "\t" + str(As) + "\t" + \
		str(Xs) + "\t" + str(Ds) + "\t" + str(Total) + "\t" + "Filler is END END END \n"
	# print T1line
	return T1line
#

if __name__ == '__main__':
    args = docopt(__doc__, version=__file__ +" " + __version__ )
    #D  print(args)


#
#################### Menu Land ####################
#

if args['--2xlsx'] or args['-f']:
	print "You are at the csv2xlsx routine -"
	print args['<ifile>']+"  " + args['<ofile>']
	print "This routine is operational (02/10/2016 03:32:42 PM)"
	#T 
	goesinta =  args['<ifile>']
	goesouta =  args['<ofile>']
	# print goesinta + "\t" + goesouta
	csv2xlsx(goesinta,goesouta)
	# 
elif args['--testcsv'] or args['-c']:
	print "You are at the testcsv routine -"
	print "This routine is operational (02/10/2016 11:58:15 AM)."
	#T 
	goesinta =  args['<ifile>']
	poplst()
	lneval(goesinta,"comma")
	# 
elif args['--testtab'] or args['-t']:
	print "You are at the testcsv routine -"
	print "This routine is operational (02/10/2016 12:38:59 PM )."
	goesinta =  args['<ifile>']
	poplst()
	lneval(goesinta,"tab")
	# 
elif args['--tespace'] or args['-s']:
	print "You are at the tespace routine -"
	print args['<ifile>']+"  " + args['<logfile>']
	print "This routine is not yet operational"
elif args['--formh&f'] or args['-a']:
	print "You are at the formh&f routine -"
	print args['<ifile>']+"  " + args['<ofile>']
	print "tab is " + str(args['--tab']) +"  csv is "  + str(args['--csv'])
	if args['--tab']: symbol = "tab"
	if args['--csv']: symbol = "csv"
	print "This routine is operational (02/11/2016 01:39:08 PM )."
	fhandf(args['<ifile>'],args['<ofile>'],symbol)
	#
elif args['--transmit'] or args['-m']:
	print "You are at the transmit routine -"
	print args['<ifile>']
	# Popen(['xterm', '-e', 'sleep 3s'])
	subprocess.call('ls | wc -l', shell=True)
	# subprocess.call('bin/hw.sh',  shell=True)
	# subprocess.call(['bin/hw.sh', 'flint'])
	# subprocess.call(['bin/sf.sh', args['<ifile>']])
	subprocess.call([comprog, args['<ifile>']])
	#
elif args['--details']:
    ''' Prints the main story about this program.'''
    os.system('cls' if os.name == 'nt' else 'clear')
    print "You are at the detail routine -"
    print "\tWelcome to the "+__appname__+" Version: "+__version__  
    print 'This is %s documentation: A general tool for NSLC support' %  os.path.basename(__file__)
    print 'This tool is currently in the %s stage, at version %s' % (__status__, __version__)
    print 'This code, produced by %s, is a work product produced for hire. It is' % __maintainer__
    print '%s, and licensed under the %s ' % ( __copyright__, __license__ )
    print "for operating this tool see commands below:"
    print "SUMMARY USAGE:"
    print "\t %s -h" % __file__
    print "DETAILED USAGE:"  
    print "\t%s  -f  or  --2xlsx    <ifile> <ofile>" %__file__
    print "   Converts CSV output of Goddard SIS NSLC to a Microsoft Excel XLSX format.\t(Works)"  
    print "\t%s   -c  or  --testcsv  <ifile> " %__file__
    print "   Tests the CSV output of Goddard SIS NSLC to conform with NSLC requirements. \t(Works)"  
    print "\t%s   -t  or  --testtab  <ifile> " %__file__
    print "   Tests if a TAB seperated file conforms with NSLC requirements.\t\t(Works)" 
    print "\t%s   -s  or  --tespace  <ifile> " %__file__
    print "   Tests if a fixed field file conforms with NSLC requirements."  
    print "\t%s   -a  or  --formh&f  <ifile> <ofile>" %__file__
    print "   Inserts 'A3' header &'T1' footer in 'tab' or 'csv' to meet NSLC requirements.(Works)"  
    print "\t%s   -m  or  --transmit ) <ifile>" %__file__
    print "   Transmitts <infile> to the NSLC destination.\t\t\t\t\t(Works)"  
    print "WHERE:"
    print "\t'<ifile>' & '<ofile>' are required as fully pathed filenames."
    print "\t'<logfile>', when implemeneted, defaults to standard output."

